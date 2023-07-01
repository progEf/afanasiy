from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect

from afas.models import User, Company, Comapany_client_id, shops, Comapany_Cl_id_AND_shops
from manager.forms import UploadFileForm, Prie_File_Form, Diskont_File_Form
from manager.models import File, File_two, File_diskont
from order.models import order_Admin, product_admin, product_discount
from django.views.generic import TemplateView, ListView
from openpyxl import load_workbook
from datetime import datetime


def meneger_cllient(request):
    a = order_Admin.objects.all()
    mas = []
    for i in a:
        ij = str(i)
        user = ij.find('User ')
        product = ij[:user]  # Название : Пиво. Вид упаковки bottle Кол-во 200; пользователя-компания-магазин
        comp_cl = ij.find('Comp_Cl_id')
        use = ij[
              comp_cl:]  # Comp_Cl_id : User : 1234, Company : ООО партнер, Shop : Shop_name : пивнуха ; Дата заказа 2023-06-04 19:32:15.268628+00:00
        usi = use.find(':')
        usp = use[usi + 1:]
        sum = product + usp  # Название : Пиво. Вид упаковки bottle Кол-во 213; пользователя-компания-магазин User : admin, Company : Butik, Shop : Shop_name : Рок ; Дата заказа 2023-06-04 14:45:16.445863+00:00
        sum_1 = sum.find(' Shop_name ')
        sum_2 = sum[:sum_1]  # sumirovat
        sum_3 = sum[sum_1:]
        sum_4 = sum_3.find(':')
        plus = sum_3[sum_4 + 1:]
        plus_1 = sum_2 + plus

        mas.append(plus_1)
    return render(request, 'manager/home.html', {'all': mas[::-1],
                                                 'user': request.user})


# Название : Пиво. Вид упаковки bottle Кол-во 200; пользователя-компания-магазин User : 1234, Company : User : 1234, Company : ООО партнер, Shop : Comp_Cl_id : User : 1234, Company : ООО партнер, Shop : Shop_name : пивнуха ; Дата заказа 2023-06-04 19:32:15.268628+00:00

# Create your views here.

class SearchResultsView(ListView):
    model = order_Admin
    template_name = 'manager/home.html'

    def get_queryset(self):  # новый
        query = self.request.GET.get('q')
        object_list = order_Admin.objects.filter(
            Q(us__id=query)
        )
        return object_list

def imports_file(request):
    if request.method == 'POST':
        if 'update' in request.POST:

            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                data = form.cleaned_data
                field = data['name']
                file_obj = File.objects.get(name=field)  # получаем объект модели File
                file_path = file_obj.file.path  # получаем путь к файлу
                with open(file_path, 'rb') as f:
                    # делаем что-то с файлом, например, читаем его содержимое
                    content = load_workbook(filename=f)
                    sheet = content.active

                    mas= [] # В конце надо что бы массив удалялся
                    for a in sheet['A']:
                        mas.append(a.value)
                    len_mas = len(mas)
                    df1 = content['Лист1']
                    for i in range(2, len_mas+ 1):
                        ma = str(df1['A' + str(i)].value), str(df1['B' + str(i)].value), str(df1['D' + str(i)].value), str(df1['E' + str(i)].value),  str(df1['G' + str(i)].value)
                        creat_user = User.objects.create_user(username=ma[0], password=ma[1]).id
                        creat_Company = Company.objects.create(company_name = ma[2], address = ma[3]).id
                        creat_comp_client = Comapany_client_id.objects.create(client_id =creat_user,comp_id =  creat_Company ).id
                        creat_shops_comp = shops.objects.create(comp_and_user_id = creat_comp_client, shops_name = ma[4]).id
                        creat_comp_shop = Comapany_Cl_id_AND_shops.objects.create(Comapany_Cl_id_id =creat_comp_client , Shops_id_id = creat_shops_comp, user_ids_id =creat_user )

                return HttpResponse("<h1>Клиенты добавлины</h1>")

        if 'price' in request.POST:
            form_2 = Prie_File_Form(request.POST, request.FILES)
            if form_2.is_valid():
                form_2.save()
                data = form_2.cleaned_data
                field = data['name']
                file_obj = File_two.objects.get(name=field)  # получаем объект модели File
                file_path = file_obj.file.path  # получаем путь к файлу
                with open(file_path, 'rb') as f:
                    # делаем что-то с файлом, например, читаем его содержимое
                    content = load_workbook(filename=f)
                    sheet = content.active

                    mas= [] # В конце надо что бы массив удалялся
                    for a in sheet['A']:
                        mas.append(a.value)
                    len_mas = len(mas)
                    df1 = content['Лист1']
                    for i in range(2, len_mas+ 1):
                        ma = str(df1['A' + str(i)].value), str(df1['B' + str(i)].value), str(df1['C' + str(i)].value)
                        find_product_admin = product_admin.objects.filter(product_name=ma[0])
                        print(len(find_product_admin))
                        if len(find_product_admin) == 0:
                            creat_product = product_admin.objects.create(product_name = ma[0], el = ma[1],price = ma[2] )

                    file_obj.delete()
                    return HttpResponse("<h1>Продукты добавлины</h1>")
    if 'diskont' in request.POST:
        form_3 = Diskont_File_Form(request.POST, request.FILES)
        if form_3.is_valid():
            form_3.save()
            data = form_3.cleaned_data
            field = data['name']
            file_obj = File_diskont.objects.get(name=field)  # получаем объект модели File
            print(file_obj)
            print(type(file_obj))
            file_path = file_obj.file.path  # получаем путь к файлу
            print(file_path)
            print(type(file_path))
            with open(file_path, 'rb') as f:
                # делаем что-то с файлом, например, читаем его содержимое
                content = load_workbook(filename=f)
                sheet = content.active

                mas = []  # В конце надо что бы массив удалялся
                for a in sheet['A']:
                    mas.append(a.value)
                len_mas = len(mas)
                df1 = content['Лист1']
                for i in range(2, len_mas + 1):
                    ma = str(df1['A' + str(i)].value), str(df1['B' + str(i)].value), str(df1['C' + str(i)].value), str(df1['D' + str(i)].value), str(df1['E' + str(i)].value)
                    filter_product = product_discount.objects.filter(product_name=ma[0], stat_time =ma[2], end_time= ma[3], price=ma[4])
                    if len(filter_product) == 0:

                        creat_product = product_discount.objects.create(product_name=ma[0], el=ma[1],stat_time =ma[2],end_time= ma[3],   price=ma[4])
                return HttpResponse("<h1>Продукты со скидкой добавлины</h1>")






    else:
        form = UploadFileForm()
        form_2 = Prie_File_Form()
        form_3 = Diskont_File_Form()
        now = datetime.now()
        formatted_date = now.strftime('%Y-%m-%d')   # Рил тайм
        last_record = File.objects.filter().order_by('-id').values_list('created_at', flat=True).first()
#        date = last_record.strftime('%Y-%m-%d')     # Дата файла


        return render(request, 'manager/imp.html', {'form': form,
                                                    'form_2': form_2,
                                                    'form_3': form_3

                                                    })


#Qweasdzxc123Qwe
#a = User.objects.create_user(username = 'ef', password = 'Qweasdzxc123Qwe', email = 'qwdasdd@gmail.com', phone_user= '21412312312')


# <form action="{% url 'search_results' %}" method="get">
#  <input name="q" type="text" placeholder="Search...">
# </form>
