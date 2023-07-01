from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from afas.models import User, Company, Comapany_client_id, Comapany_Cl_id_AND_shops, shops
from manager.models import File, File_two, File_diskont
from order.forms import UserCompShop, Product_Amout, ProductAdminForm, Product_order_user, update_ord_form, \
    choice_company
from order.models import User_Comp_Shop, product_admin, order_products_user, order_Admin, product_discount, \
    order_companu, key_User_Comp_Shop
# from django.utils.timezone import utc
from datetime import datetime
from openpyxl import load_workbook
import os


def firma_and_order312(request):
    if request.method == 'POST':
        form = UserCompShop(request.POST, user=request.user)

        if form.is_valid():
            data_1 = form.cleaned_data
            # field_1 = data_1['Comapany_Cl_id'].id
            field_shop = data_1['Shops_id'].id
            us = User.objects.get(username=request.user.username)  # Узнаем usera сессии     ОСТАВИТЬ
            a = User.objects.filter(username=us).values_list('pk', flat=True)
            find_pifye = User_Comp_Shop.objects.filter(client_id_id=a[0], Shops_id_id=field_shop)

            try:
                print(find_pifye)
            except:
                print('asdafafasd')

            find_comp = User_Comp_Shop.objects.filter(client_id_id=a[0], Shops_id_id=field_shop)
            try:
                a_find = find_comp[0]
                if a_find != None:
                    pifye = User_Comp_Shop.objects.filter(client_id_id=a[0], Shops_id_id=field_shop).values_list('id',
                                                                                                                 flat=True)
                    print(pifye)
                    id_User_Comp_Shop = User_Comp_Shop.objects.filter(pk=pifye[0]).values_list('Shops_id', flat=True)
                    find_Comapany_Cl_id_AND_shops = User_Comp_Shop.objects.filter(Shops_id_id=id_User_Comp_Shop[0])
                else:
                    pifye = User_Comp_Shop.objects.create(client_id_id=a, Shops_id_id=field_shop).id
                    id_User_Comp_Shop = User_Comp_Shop.objects.filter(pk=pifye).values_list('Shops_id', flat=True)
                    find_Comapany_Cl_id_AND_shops = User_Comp_Shop.objects.filter(Shops_id_id=id_User_Comp_Shop[0])

            except:
                pifye = User_Comp_Shop.objects.create(client_id_id=a, Shops_id_id=field_shop).id
                id_User_Comp_Shop = User_Comp_Shop.objects.filter(pk=pifye).values_list('Shops_id', flat=True)
                find_Comapany_Cl_id_AND_shops = User_Comp_Shop.objects.filter(Shops_id_id=id_User_Comp_Shop[0])

            if len(find_Comapany_Cl_id_AND_shops) >= 1:
                procent = User_Comp_Shop.objects.filter(Shops_id_id=id_User_Comp_Shop[0]).values_list('status_procent',
                                                                                                      flat=True)
                try:
                    procent_dis = procent[0]

                    # find_Comapany_Cl_id_AND_shops.delete()
                    # find_comp = User_Comp_Shop.objects.filter(client_id_id=a, Shops_id_id=field_shop, status_procent=procent_dis)
                    # print(find_comp[0])
                    # pifye = User_Comp_Shop.objects.create(client_id_id=a, Shops_id_id=field_shop, status_procent = procent_dis).id

                except:
                    # find_Comapany_Cl_id_AND_shops.delete()

                    find_comp = User_Comp_Shop.objects.filter(client_id_id=a, Shops_id_id=field_shop).values_list('id',
                                                                                                                  flat=True)

                    if len(find_comp) >= 1:
                        pifye = User_Comp_Shop.objects.filter(id=find_comp[0]).update(client_id_id=a,
                                                                                      Shops_id_id=field_shop).id
                    # pifye = User_Comp_Shop.objects.create(client_id_id=a, Shops_id_id=field_shop).id

            #################
            id_user = key_User_Comp_Shop.objects.filter(client_id_user_id=a[0]).values_list('client_id_user', flat=True)
            try:
                id_user_id = id_user[0]
                filter_del = key_User_Comp_Shop.objects.filter(client_id_user_id=a[0])
                filter_del.delete()
                creat_key_User_Comp_Shop = key_User_Comp_Shop.objects.create(client_id_user_id=a, key_comp_id=pifye)
            except:
                creat_key_User_Comp_Shop = key_User_Comp_Shop.objects.create(client_id_user_id=a, key_comp_id=pifye)

            # 'order/shop/'
            return HttpResponseRedirect('http://127.0.0.1:8000/order/shop/')
    else:
        form = UserCompShop(request.POST, user=request.user)
        # product_admin_object = product_admin.objects.first()
        # from_product =ProductAdminForm(instance=product_admin_object)
        now = datetime.now()
        formatted_date = now.strftime('%Y-%m-%d')  # Рил тайм
        last_record_1 = File_diskont.objects.filter().order_by('-id').values_list('created_at', flat=True).first()

        last_record = File_two.objects.filter().order_by('-id').values_list('created_at', flat=True).first()
        try:
            date = last_record.strftime('%Y-%m-%d')  # Дата файла
            date_1 = last_record_1.strftime('%Y-%m-%d')  # Дата файла

            if formatted_date != date:
                last_name = File_two.objects.filter().order_by('-id').values_list('file', flat=True).first()
                direkt = os.path.join(os.getcwd(), last_name)
                print(type(direkt))

                # file_path = direkt.file.path
                file_path = os.path.join(direkt)
                with open(file_path, 'rb') as f:
                    #
                    # делаем что-то с файлом, например, читаем его содержимое
                    content = load_workbook(filename=f)
                    sheet = content.active
                    mas = []  # В конце надо что бы массив удалялся
                    for a in sheet['A']:
                        mas.append(a.value)
                    len_mas = len(mas)
                    df1 = content['Лист1']
                    for i in range(2, len_mas + 1):
                        ma = str(df1['A' + str(i)].value), str(df1['B' + str(i)].value), str(df1['C' + str(i)].value)
                        find_product_admin = product_admin.objects.filter(product_name=ma[0])
                        print(len(find_product_admin))
                        print('sadasfasg')
                        if len(find_product_admin) == 0:
                            creat_product = product_admin.objects.create(product_name=ma[0], el=ma[1], price=ma[2])
                            print('asdsadsadds')

            if formatted_date != date_1:
                last_name = File_diskont.objects.filter().order_by('-id').values_list('file', flat=True).first()
                direkt = os.path.join(os.getcwd(), last_name)
                print(type(direkt))

            # file_path = direkt.file.path
                file_path = os.path.join(direkt)
                with open(file_path, 'rb') as f:
                #
                # делаем что-то с файлом, например, читаем его содержимое
                    content = load_workbook(filename=f)
                    sheet = content.active
                    mas = []  # В конце надо что бы массив удалялся
                    for a in sheet['A']:
                        mas.append(a.value)
                    len_mas = len(mas)
                    df1 = content['Лист1']
                    for i in range(2, len_mas + 1):
                        ma = str(df1['A' + str(i)].value), str(df1['B' + str(i)].value), str(df1['C' + str(i)].value), str(
                            df1['D' + str(i)].value), str(df1['E' + str(i)].value)
                        filter_product = product_discount.objects.filter(product_name=ma[0], stat_time=ma[2],
                                                                     end_time=ma[3], price=ma[4])
                        print('sdasfaasfw')
                        if len(filter_product) == 0:
                            creat_product = product_discount.objects.create(product_name=ma[0], el=ma[1], stat_time=ma[2],
                                                                        end_time=ma[3], price=ma[4])
                            print('sadasfasg')

        except:
            print('Ошибка в order/views в проверке файла exls')
        return render(request, 'orders.html', {
            'form_shop_comp': form,
            # 'from_products': from_product,
        })


def order_user(request):
    if request.method == 'POST':
        from_order = Product_Amout(request.POST, user=request.user)
        us = User.objects.get(username=request.user.username)  # Узнаем usera сессии     ОСТАВИТЬ
        us_1 = User.objects.get(username=request.user.username).id  # Узнаем usera сессии     ОСТАВИТЬ

        orde = order_products_user.objects.filter(id_us=us)

        if from_order.is_valid():
            print('asdasfasfasfdddd')

            from_order.save()
            ID_user_c_s = from_order.cleaned_data['id_user'].id
            amaut = from_order.cleaned_data['amount']
            ucs = order_products_user.objects.filter(id_user_id=ID_user_c_s).values_list('id_user', flat=True)

            us_ucs = User_Comp_Shop.objects.filter(id=ucs[0]).values_list('status_procent', flat=True)

            stat = us_ucs[0]

            try:  # Товары без скидки
                ID_product_admin = from_order.cleaned_data['id_product'].id
                b = product_admin.objects.filter(pk=ID_product_admin).values_list('price', flat=True)
                b1 = b[0]
                status_dis = order_products_user.objects.filter(id_us_id=us_1, id_user_id=ID_user_c_s,
                                                                amount=amaut).values_list('id_product',
                                                                                          flat=True)

                if stat == None or stat == 0:

                    c = order_products_user.objects.filter(id_us_id=us_1, id_product_id=ID_product_admin,
                                                           amount=amaut).update(sumir=amaut * b1)

                    kol_bo = order_products_user.objects.filter(id_us_id=us_1, id_user_id=ID_user_c_s,
                                                                id_product_id=ID_product_admin).values_list('pk',
                                                                                                            flat=True)
                    len_kol = len(kol_bo)
                    if len_kol > 1:
                        mas_pk_dis = kol_bo
                        for i in mas_pk_dis:
                            pk_ord_user = order_products_user.objects.filter(pk=i)
                            pk_ord_user.delete()
                        return HttpResponse("<h1>Нельзя добавлять два одинаковых товара</h1>")

                else:
                    c_1 = order_products_user.objects.filter(id_us_id=us_1, id_product_id=ID_product_admin,
                                                             amount=amaut).update(
                        sumir=(amaut * b1))
                    #   sumir=(amaut * b1) - ((amaut * b1) * (stat / 100)))

                    kol_bo = order_products_user.objects.filter(id_us_id=us_1, id_user_id=ID_user_c_s,
                                                                id_product_id=ID_product_admin).values_list('pk',
                                                                                                            flat=True)
                    len_kol = len(kol_bo)
                    if len_kol > 1:
                        mas_pk_dis = kol_bo
                        for i in mas_pk_dis:
                            pk_ord_user = order_products_user.objects.filter(pk=i)
                            pk_ord_user.delete()
                        return HttpResponse("<h1>Нельзя добавлять два одинаковых товара</h1>")

            except AttributeError:  # Товары со скидкой
                ID_product_admin = from_order.cleaned_data['id_product_discont'].id
                b = product_discount.objects.filter(pk=ID_product_admin).values_list('price', flat=True)

                b1 = b[0]  # Цена
                status_dis = order_products_user.objects.filter(id_us_id=us_1, id_user_id=ID_user_c_s,
                                                                amount=amaut).values_list('id_product_discont',
                                                                                          flat=True)

                if stat == None or stat == 0:

                    c = order_products_user.objects.filter(id_us_id=us_1, id_product_discont_id=ID_product_admin,
                                                           amount=amaut).update(
                        sumir=amaut * b1 - ((amaut * b1) * (stat / 100)))
                    kol_bo = order_products_user.objects.filter(id_us_id=us_1, id_user_id=ID_user_c_s,
                                                                id_product_discont_id=ID_product_admin).values_list(
                        'pk', flat=True)

                    len_kol = len(kol_bo)
                    if len_kol > 1:
                        mas_pk_dis = kol_bo
                        for i in mas_pk_dis:
                            pk_ord_user = order_products_user.objects.filter(pk=i)
                            pk_ord_user.delete()
                        return HttpResponse("<h1>Нельзя добавлять два одинаковых товара</h1>")
                else:  # Если есть общая скидка
                    c_1 = order_products_user.objects.filter(id_us_id=us_1, id_product_discont=ID_product_admin,
                                                             amount=amaut).update(
                        sumir=amaut * b1 - ((amaut * b1) * ((stat / 100))))

                    kol_bo = order_products_user.objects.filter(id_us_id=us_1, id_user_id=ID_user_c_s,
                                                                id_product_discont_id=ID_product_admin).values_list(
                        'pk', flat=True)
                    len_kol = len(kol_bo)
                    if len_kol > 1:
                        mas_pk_dis = kol_bo
                        for i in mas_pk_dis:
                            pk_ord_user = order_products_user.objects.filter(pk=i)
                            pk_ord_user.delete()
                        return HttpResponse("<h1>Нельзя добавлять два одинаковых товара </h1>")
            ucs = order_products_user.objects.filter(id_user_id=ID_user_c_s).values_list('id_user', flat=True)

            us_ucs = User_Comp_Shop.objects.filter(id=ucs[0]).values_list('status_procent', flat=True)

            stat = us_ucs[0]
            # status_dis = order_products_user.objects.filter(id_us_id=us_1, id_user_id=ID_user_c_s, amount=amaut).values_list('id_product_discont',
            #                                                         flat=True)  # Унаем йд статуса скидки на продукт
            # status_dis_key = status_dis[0] # None or pk

            return render(request, 'orders_shop.html', {
                'from_or': from_order,
                'ord': orde,
                'pric': us_ucs

                # 'ord': orde
            })
    else:
        from_order = Product_Amout(request.POST, user=request.user)
        us = User.objects.get(username=request.user.username)  # Узнаем usera сессии     ОСТАВИТЬ
        us_id = User.objects.get(username=request.user.username).id
        client_id = User.objects.filter(username=us).values_list('id', flat=True).get()

        orde = order_products_user.objects.filter(id_us=us)
        orsd = order_products_user.objects.filter(id_us=us).values_list('amount',
                                                                        flat=True)  # Колво выбранного товара <QuerySet [200, 213, 213, 200, 200, 3]>
        b = order_products_user.objects.filter(id_us=us).values_list('id_product',
                                                                     flat=True)  # Получаем id продуктов привязаной к use

        user_ids = order_products_user.objects.filter(id_us_id=us_id).values_list('id_user', flat=True)
        try:
            user_ids_m = user_ids[0]
        except:
            user_ids_m = 0
        # print(user_ids)
        us_ucs = User_Comp_Shop.objects.filter(id=user_ids_m).values_list('status_procent', flat=True)

        # now = datetime.datetime.now().strftime('%H:%M:%S')
        now = datetime.now()
        formatted_date = now.strftime('%Y-%m-%d')
        id_product_discount = product_discount.objects.all().values_list('pk', flat=True)
        for ids in id_product_discount:
            one_id = product_discount.objects.filter(pk=ids).values_list('end_time', flat=True)
            ones_id = str(one_id[0])
            if ones_id == formatted_date:
                product_name_dis = product_discount.objects.filter(pk=ids).values_list('product_name', flat=True)
                product_name_dis_tr = product_name_dis[0]  # product_name
                el_dis = product_discount.objects.filter(pk=ids).values_list('el', flat=True)
                el_dis_tr = el_dis[0]  # ед измерения
                price_dis = product_discount.objects.filter(pk=ids).values_list('price', flat=True)
                price_dis_tr = price_dis[0]
                product_admin.objects.create(product_name=product_name_dis_tr, el=el_dis_tr, price=price_dis_tr)
                id_dis = product_discount.objects.filter(pk=ids)  # .values_list('product_name', flat=True)
                id_dis.delete()

        status = key_User_Comp_Shop.objects.filter(client_id_user_id=us_id).values_list('key_comp', flat=True)
        status_dis = User_Comp_Shop.objects.filter(id=status[0]).values_list('status_procent', flat=True)

        return render(request, 'orders_shop.html', {
            'from_or': from_order,
            'ord': orde,
            'pric': us_ucs,
            'time': formatted_date
        })


def order(request):
    if request.method == 'POST':

        from_order = Product_order_user(request.POST, user=request.user)
        if from_order.is_valid():
            from_data = from_order.cleaned_data['data_time']  # Узнаем дату
            from_user_comp = from_order.cleaned_data['user_id']  # Узнаем id фирмы

            us = User.objects.get(username=request.user.username)  # Узнаем usera сессии     ОСТАВИТЬ
            us_1 = User.objects.get(username=us).id  # Узнаем id usera сессии     ОСТАВИТЬ
            us_2 = User_Comp_Shop.objects.filter(client_id=us_1).first()
            a = order_products_user.objects.filter(id_us=us, id_user_id=from_user_comp).values_list('pk',
                                                                                                    flat=True)  # id ячеек в табл order_products_user
            for value in a:
                c1 = order_products_user.objects.get(pk=value)
                c2 = order_products_user.objects.filter(pk=value).values_list('id_product_discont', flat=True)
                prouct_no_dis = order_products_user.objects.filter(pk=value).values_list('id_product', flat=True)
                prouct_no_dis_one = prouct_no_dis[0]  # узнаем id продуктов без скидкой
                if prouct_no_dis_one != None:
                    order_Admin.objects.create(us=c1, user_id=us_2, data_time=from_data)

                print(prouct_no_dis_one)
                c2_0 = c2[0]  # узнаем id продуктов со скидкой
                product_discount_usera = product_discount.objects.filter(pk=c2_0).values_list('end_time', flat=True)
                # print(product_discount_usera)

                try:
                    product_discount_usera_time = product_discount_usera[0]  # Узнаем конец срока акции

                    if product_discount_usera_time >= from_data:
                        try:
                            order_Admin.objects.create(us=c1, user_id=us_2, data_time=from_data)
                        except:
                            print('Херова')
                    else:
                        fin_time_dis = product_discount.objects.filter(
                            end_time=product_discount_usera_time).values_list('discount', flat=True)
                        fin_time_dis_one = fin_time_dis[0]  # Узнаем скидку товара
                        # print(fin_time_dis_one)
                        sum_pro_dis = order_products_user.objects.filter(id_product_discont_id=c2_0,
                                                                         id_us_id=us_1).values_list('sumir', flat=True)
                        sum_pro_dis_one = sum_pro_dis[0]  # Сумма товара
                        # print(sum_pro_dis_one)

                        sum_pro_dis = order_products_user.objects.filter(id_product_discont_id=c2_0,
                                                                         id_us_id=us_1).update(
                            sumir=sum_pro_dis_one + (sum_pro_dis_one * (fin_time_dis_one / 100)))



                except:
                    pass
                orde = order_products_user.objects.filter(id_us_id=us, id_user_id=from_user_comp)

            # delet_all_usre =
            return render(request, 'user_order.html', {
                'ord': orde,
                # 'q': q,
                'form': from_order

            })
    else:
        from_order = Product_order_user(request.POST, user=request.user)

        us = User.objects.get(username=request.user.username)
        us_id = User.objects.get(username=request.user.username).id  # Узнаем usera сессии     ОСТАВИТЬ

        # Узнаем usera сессии     ОСТАВИТЬ
        find_user_com = key_User_Comp_Shop.objects.filter(client_id_user_id=us_id).values_list('key_comp', flat=True)

        orde = order_products_user.objects.filter(id_us=us, id_user_id=find_user_com[0])
        return render(request, 'user_order.html', {
            'ord': orde,
            'form': from_order
        })


def update_ord(request):  # Заказ

    if request.method == 'POST':
        form = update_ord_form(request.POST, user=request.user)

        us_1 = User.objects.get(username=request.user.username).id  # Узнаем usera сессии     ОСТАВИТЬ
        form_ch = choice_company(request.POST, user=request.user)

        if 'add' in request.POST:
            if form_ch.is_valid():
                form_ch_id_user = form_ch.cleaned_data['id_comp'].id  # Узнаем магазин
                a = order_companu.objects.filter(id_user_id=us_1)
                b = len(a)
                if b > 0:
                    a.delete()
                    add_cp = order_companu.objects.create(id_user_id=us_1, id_comp_id=form_ch_id_user)
                    add_cp.save()
                else:
                    add_cp = order_companu.objects.create(id_user_id=us_1, id_comp_id=form_ch_id_user)
                    add_cp.save()

        return HttpResponseRedirect('http://127.0.0.1:8000/up_ord/chj/')

    else:
        form = update_ord_form(request.POST, user=request.user)
        form_ch = choice_company(request.POST, user=request.user)

        return render(request, 'update_ord/up_ord.html', {  # 'form': form,
            'form_ch': form_ch

        })


def rog_ord(request):
    if request.method == 'POST':
        form = update_ord_form(request.POST, user=request.user)

        us_1 = User.objects.get(username=request.user.username).id  # Узнаем usera сессии     ОСТАВИТЬ
        form_ch = choice_company(request.POST, user=request.user)

        if 'del_all' in request.POST:
            if form.is_valid():
                from_id_user = form.cleaned_data['id_user'].id  # Узнаем магазин

                prod = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_user)
                prod.delete()
                return HttpResponseRedirect('http://127.0.0.1:8000/up_ord/chj/')

        if 'del' in request.POST:
            form = update_ord_form(request.POST, user=request.user)

            if form.is_valid():
                from_amount = form.cleaned_data['amount']  # Узнаем кол_во
                from_id_user = form.cleaned_data['id_user'].id  # Узнаем магазин
                try:  # если не None
                    from_id_product = form.cleaned_data['id_product'].id  # Узнаем id продукта без скидкой
                except:
                    from_id_product = form.cleaned_data['id_product']  # Узнаем id продукта без скидкой

                try:  # если не None

                    from_id_product_discont = form.cleaned_data[
                        'id_product_discont'].id  # Узнаем id продукта со скидкой
                except:
                    from_id_product_discont = form.cleaned_data['id_product_discont']

                ##############################################
                qwe = from_id_product
                if qwe:

                    prod = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_user,
                                                              id_product_id=from_id_product
                                                              )
                    yes_or_no_len = len(prod)
                    if yes_or_no_len != 0:
                        prod.delete()
                    else:
                        return HttpResponse("<h1>У вас не было этого товара в корзине</h1>")

                else:
                    prod_dis = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_user,
                                                                  id_product_discont_id=from_id_product_discont)
                    yes_or_no_len = len(prod_dis)
                    if yes_or_no_len != 0:
                        prod_dis.delete()
                    else:
                        return HttpResponse("<h1>У вас не было этого товара в корзине</h1>")

        if 'update' in request.POST:
            form = update_ord_form(request.POST, user=request.user)

            if form.is_valid():
                from_amount = form.cleaned_data['amount']  # Узнаем кол_во
                from_id_user = form.cleaned_data['id_user'].id  # Узнаем магазин
                try:  # если не None
                    from_id_product = form.cleaned_data['id_product'].id  # Узнаем id продукта без скидкой
                except:
                    from_id_product = form.cleaned_data['id_product']  # Узнаем id продукта без скидкой

                try:  # если не None

                    from_id_product_discont = form.cleaned_data[
                        'id_product_discont'].id  # Узнаем id продукта со скидкой
                except:
                    from_id_product_discont = form.cleaned_data['id_product_discont']  # Узнаем id продукта со скидкой

                ####### Знаем статус магазина ################

                ucs = order_products_user.objects.filter(id_user_id=from_id_user).values_list('id_user', flat=True)
                us_ucs = User_Comp_Shop.objects.filter(id=ucs[0]).values_list('status_procent', flat=True)
                stat = us_ucs[0]  # int or None

                ##############################################

                # Товары без скидки
                b = product_admin.objects.filter(pk=from_id_product).values_list('price', flat=True)
                if b:
                    b1 = b[0]
                    status_dis = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_product,
                                                                    amount=from_amount).values_list('id_product',
                                                                                                    flat=True)
                    if stat == None or stat == 0:
                        yes_or_no = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_user,
                                                                       id_product_id=from_id_product)
                        yes_or_no_len = len(yes_or_no)
                        if yes_or_no_len != 0:
                            c = order_products_user.objects.filter(id_us_id=us_1, id_product_id=from_id_product,
                                                                   id_user_id=from_id_user).update(
                                sumir=from_amount * b1,
                                amount=from_amount)
                            q12 = order_products_user.objects.filter(id_us_id=us_1, id_product_id=from_id_product,
                                                                     )
                        else:
                            return HttpResponse("<h1>У вас нету этого товара в корзине</h1>")

                    else:
                        yes_or_no = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_user,
                                                                       id_product_id=from_id_product)
                        yes_or_no_len = len(yes_or_no)
                        if yes_or_no_len != 0:
                            c_1 = order_products_user.objects.filter(id_us_id=us_1, id_product_id=from_id_product,
                                                                     id_user_id=from_id_user
                                                                     ).update(
                                sumir=(from_amount * b1), amount=from_amount)

                        else:
                            return HttpResponse("<h1>У вас нету этого товара в корзине</h1>")
                # Товары со скидкой

                else:
                    b = product_discount.objects.filter(pk=from_id_product_discont).values_list('price', flat=True)

                    b1 = b[0]  # Цена

                    ucs = order_products_user.objects.filter(id_user_id=from_id_user).values_list('id_user', flat=True)
                    us_ucs = User_Comp_Shop.objects.filter(id=ucs[0]).values_list('status_procent', flat=True)
                    stat = us_ucs[0]  # int or None
                    if stat == None or stat == 0:
                        print('asd')

                        c_3214 = order_products_user.objects.filter(id_us_id=us_1,
                                                                    id_product_discont_id=from_id_product_discont,
                                                                    )
                        yes_or_no = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_user,
                                                                       id_product_discont_id=from_id_product_discont)
                        yes_or_no_len = len(yes_or_no)
                        if yes_or_no_len != 0:

                            c = order_products_user.objects.filter(id_us_id=us_1,
                                                                   id_product_discont_id=from_id_product_discont,
                                                                   id_user_id=from_id_user).update(amount=from_amount,
                                                                                                   sumir=from_amount * b1)
                        else:
                            return HttpResponse("<h1>У вас нету этого товара в корзине</h1>")
                    else:  # Если есть общая скидка
                        print('sagfasfaf')
                        yes_or_no = order_products_user.objects.filter(id_us_id=us_1, id_user_id=from_id_user,
                                                                       id_product_discont_id=from_id_product_discont)
                        yes_or_no_len = len(yes_or_no)
                        if yes_or_no_len != 0:

                            c_1 = order_products_user.objects.filter(id_us_id=us_1,
                                                                     id_product_discont=from_id_product_discont,
                                                                     id_user_id=from_id_user).update(
                                sumir=from_amount * b1 - ((from_amount * b1) * ((stat / 100))),
                                amount=from_amount)
                        else:
                            return HttpResponse("<h1>У вас нету этого товара в корзине</h1>")

        return render(request, 'update_ord/chos_ord.html', {'form': form})

    else:
        form = update_ord_form(request.POST, user=request.user)
        form_ch = choice_company(request.POST, user=request.user)

        return render(request, 'update_ord/chos_ord.html', {  # 'form': form,
            'form': form

        })


def update_basket(request):  # Корзина
    return render(request, 'update_ord/up_basket.html', {})
